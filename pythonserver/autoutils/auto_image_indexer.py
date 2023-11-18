import os
import time
from datetime import datetime
from PIL import Image
from openpyxl import load_workbook, Workbook

def process_images():
    folder_path = '../images'  # Replace with your folder path
    excel_file = '../image_data.xlsx'

    # Load or initialize the workbook
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['id', 'file', 'width', 'height', 'size', 'format', 'type', 'subtype', 'tags', 'votes', 'active', 'timestamp'])

    # Remove inactive images and their entries
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[10].value == 0:  # Column K corresponds to 'active'
            file_name = row[1].value  # Column B corresponds to 'file'
            file_path = os.path.join(folder_path, file_name)
            if os.path.exists(file_path):
                os.remove(file_path)  # Delete the file
                print(f"Deleted file: {file_name}")
            ws.delete_rows(row[0].row)  # Delete the row

    # Save changes after removing inactive images
    wb.save(excel_file)

    # Process new images
    file_names = os.listdir(folder_path)
    seen_files = {row[1] for row in ws.iter_rows(min_row=2, max_col=2, values_only=True)}
    next_id = max((row[0] for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if isinstance(row[0], int)), default=0) + 1

    for file_name in file_names:
        if file_name not in seen_files and file_name.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
            file_path = os.path.join(folder_path, file_name)
            try:
                with Image.open(file_path) as img:
                    width, height = img.size
                    file_size = os.path.getsize(file_path)
                    image_format = img.format
                    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    ws.append([next_id, file_name, width, height, file_size, image_format, 'image', '', '', 0, 1, timestamp])
                    next_id += 1
                    print(f'Indexed file: {file_name}')
            except IOError:
                print(f'Cannot open file: {file_name}')

    # Save the workbook after updating
    wb.save(excel_file)
    print("Processing complete.")

# Run the process in an infinite loop
while True:
    process_images()
    print("Waiting for next run...")
    time.sleep(120)  # Wait for 120 seconds (2 minutes)
