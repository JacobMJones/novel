from datetime import datetime
import string
import openpyxl as px
import ast

def preprocess_text(text: str) -> str:
    """Convert text to lowercase and remove punctuation."""
    return text.lower().translate(str.maketrans("", "", string.punctuation))

def get_next_id(sheet):
    """Get the next ID based on the highest ID in the ID field or the number of rows."""
    last_row = sheet.max_row
    print('last', last_row)
    if last_row > 1:
        # Fetch the ID from the last row and increment it by 1
        last_id = sheet.cell(row=last_row, column=1).value
        try:
            next_id = int(last_id) + 1
        except (ValueError, TypeError):
            # If the last ID is not an integer, use the row count as ID
            next_id = last_row
    else:
        # If the sheet is empty (only headers), start from 1
        next_id = 1

    return next_id

def load_or_create_workbook(file_path: str):
    """Load an existing workbook or create a new one if it doesn't exist."""
    try:
        workbook = px.load_workbook(file_path)
        sheet = workbook.active
    except FileNotFoundError:
        # If the workbook doesn't exist, create it with the specified headers
        workbook = px.Workbook()
        sheet = workbook.active
        headers = ['ID', 'Content', 'Type', 'Subtype', 'Priority', 'Tags', 'Timestamp']
        sheet.append(headers)

    return workbook, sheet

def main():
    """Main function to append new data to the Excel file."""
    while True:  # Start an infinite loop
        try:
            # User inputs
            content_subtype = input("Enter the content subtype (or type 'exit' to quit): ").strip().lower()
            if content_subtype == 'exit':  # Check if user wants to exit
                break

            content_tags = input("Enter tags, separated by commas (or type 'exit' to quit): ").strip()
            if content_tags == 'exit':  # Check if user wants to exit
                break

            text_to_add = input("Paste the text you want to add (or type 'exit' to quit): ").strip()
            if text_to_add == 'exit':  # Check if user wants to exit
                break

            # Process the text
            processed_text = preprocess_text(text_to_add)

            # Create timestamp
            timestamp = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')

            # Load or create workbook
            file_path = 'text_data.xlsx'
            workbook, sheet = load_or_create_workbook(file_path)

            # Get the next ID
            next_id = get_next_id(sheet)

            # New entry to match Excel columns
            new_entry = [next_id, processed_text, 'text', content_subtype, 1, content_tags, timestamp]

            # Append new entry and save the workbook
            sheet.append(new_entry)
            workbook.save(file_path)

        except Exception as e:
            print(f"Error occurred: {e}")

if __name__ == "__main__":
    main()
